r"""
Extrai os dados do painel "Agenda" a partir de AGENDA.xlsx (aba "RESUMO DO
DIA") e salva um dados.json pronto pro gerador de HTML consumir.

Layout da aba (confirmado 28/08): a mesma tabela de 6 supervisores se repete
empilhada verticalmente na aba inteira — cada bloco começa numa linha de
cabeçalho (coluna C = nome do supervisor, ex: "SUP RICHARD" ou, no primeiro
bloco, "2 - LEANDRO FREITAS"; colunas D em diante = rótulos de coluna fixos),
seguida de 1 linha por RCA até a próxima linha em branco. As linhas de RCA
têm a coluna C prefixada com espaços não-quebráveis (\xa0) — é assim que dá
pra diferenciar "linha de cabeçalho" de "linha de RCA" sem depender de
formatação.

Por RCA (coluna, 1-indexed a partir de C=3):
- C (3)  = "código - nome - rota" (prefixo \xa0\xa0...)
- D (4)  = horário do 1º cliente
- E (5)  = horário do último cliente
- F (6)  = T.M.A. (tempo médio de atendimento)
- G (7)  = meta da agenda (nº de clientes agendados)
- I (9)  = visitados (realizado)
- J (10) = % da agenda (visitado/agenda)
- O (15) = dentro da agenda (visitas feitas dentro do horário agendado)
- Q (17) = % dentro da agenda
- T (20) = pedidos feitos dentro da agenda
- U (21) = conversão (pedidos-agenda / visitado)
- V (22) = total de pedidos do dia
(H, K, L, M, N, P, R, S, W, X são colunas duplicadas/auxiliares da própria
planilha — não carregam informação nova, ficam de fora.)

A linha "EQUIPE GYN" (~linha 65) fecha a aba com um resumo do time inteiro,
mas tem fórmulas quebradas (#REF!/#VALUE!) — em vez de usar ela, o time é
recalculado no próprio painel a partir dos RCAs (mesmo padrão dos outros
painéis).
"""

import datetime
import json
import os
import re

import openpyxl

CAMINHO_AGENDA = r"C:\Users\edmar\Desktop\AGENDA.xlsx"

PASTA_BASE = os.path.dirname(os.path.abspath(__file__))
CAMINHO_SAIDA = os.path.join(PASTA_BASE, "dados.json")


def _num(v):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        v = v.strip().replace(",", ".")
        try:
            return float(v)
        except ValueError:
            return 0
    return 0


def _horario(v):
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    return ""


def _nome_supervisor(texto):
    """'SUP RICHARD' -> 'RICHARD' · '2 - LEANDRO FREITAS' -> 'LEANDRO'
    (primeiro nome, pra bater com a convenção usada nos outros painéis)."""
    texto = texto.strip()
    if texto.upper().startswith("SUP "):
        nome = texto[4:].strip()
    else:
        m = re.match(r"^\d+\s*-\s*(.+)$", texto)
        nome = m.group(1).strip() if m else texto
    return nome.split()[0].upper() if nome else texto.upper()


def _parse_rca(ws, r, supervisor):
    nome_bruto = str(ws.cell(row=r, column=3).value).replace("\xa0", "").strip()
    m = re.match(r"^(\d+)\s*-\s*(.+)$", nome_bruto)
    if not m:
        return None
    codigo = m.group(1)
    partes = [p.strip() for p in m.group(2).split(" - ")]
    nome = partes[0]
    rota = " - ".join(partes[1:]) if len(partes) > 1 else ""

    return {
        "codigo": codigo,
        "nome": nome,
        "rota": rota,
        "supervisor": supervisor,
        "primeiro_cliente": _horario(ws.cell(row=r, column=4).value),
        "ultimo_cliente": _horario(ws.cell(row=r, column=5).value),
        "tma": _horario(ws.cell(row=r, column=6).value),
        "meta_agenda": int(_num(ws.cell(row=r, column=7).value)),
        "visitado": int(_num(ws.cell(row=r, column=9).value)),
        "pct_agenda": _num(ws.cell(row=r, column=10).value),
        "dentro_agenda": int(_num(ws.cell(row=r, column=15).value)),
        "pct_dentro_agenda": _num(ws.cell(row=r, column=17).value),
        "ped_agenda": int(_num(ws.cell(row=r, column=20).value)),
        "conversao": _num(ws.cell(row=r, column=21).value),
        "total_pedidos": int(_num(ws.cell(row=r, column=22).value)),
    }


def extrair():
    wb = openpyxl.load_workbook(CAMINHO_AGENDA, data_only=True)
    ws = wb["RESUMO DO DIA"]

    dia = ws["I2"].value

    rcas = []
    supervisor_atual = None
    for r in range(1, ws.max_row + 1):
        valor = ws.cell(row=r, column=3).value
        if valor is None or str(valor).strip() == "":
            continue
        texto = str(valor)
        if texto.strip().upper().startswith("EQUIPE"):
            break
        if texto.startswith("\xa0"):
            rca = _parse_rca(ws, r, supervisor_atual)
            if rca:
                rcas.append(rca)
        else:
            supervisor_atual = _nome_supervisor(texto)

    return rcas, dia


def main():
    rcas, dia = extrair()
    saida = {"rcas": rcas, "dia": dia}
    with open(CAMINHO_SAIDA, "w", encoding="utf-8") as f:
        json.dump(saida, f, ensure_ascii=False, indent=2)
    print(f"{len(rcas)} RCAs extraidos (dia {dia}). Salvo em: {CAMINHO_SAIDA}")


if __name__ == "__main__":
    main()
